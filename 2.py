import gzip
import json
import os
from queue import Queue

import openpyxl
import requests
from lxml import etree
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def get_region_and_filtvalue():
    html_info = requests.get('http://www.cpppc.org:8086/pppcentral/map/toPPPList.do')
    html = etree.HTML(html_info.text)
    region_items = html.xpath('//div[@class="v-fold"]/ul[@class="l-list1"]//a/text()')[3:-20]
    filtvalue_items = html.xpath('//div[@class="v-fold"]/ul[@class="l-list1"]//a/@filtvalue')[3:-20]
    for k, v in enumerate(region_items):
        region_items[k] = str(v).strip()
    items = dict(zip(region_items, filtvalue_items))
    return items


def get_project_details(PROJ_RID):
    url = 'http://www.cpppc.org:8083/efmisweb/ppp/projectLibrary/getProjInfoNational.do?projId=' + PROJ_RID
    html_info = requests.get(url)
    html = etree.HTML(html_info.text)
    details_items = html.xpath('//div[@class="margin"]/table//td[@colspan="5"]/text()')
    social_capital_items = html.xpath('//div[@id="con_ss_1"]/div/table[1]/tbody/tr[4]/td[2]/text()')
    ESTIMATE_COPER_items = html.xpath('//div[@id="con_ss_1"]/div/table[1]/tbody/tr[3]/td[2]/text()')
    print(ESTIMATE_COPER_items)
    return {'Project_demonstration_level': ''.join(details_items),
            'Ways_of_purchasing_social_capital': ''.join(social_capital_items),
            'ESTIMATE_COPER': ''.join(ESTIMATE_COPER_items)}


def get_project_list_by_region(distStr):
    url = 'http://www.cpppc.org:8086/pppcentral/map/getPPPList.do'
    params = {
        'queryPage': '1',
        'distStr': distStr,
        'projStateType': '1'
    }
    html_info = requests.get(url, params=params).content
    json_info = json.loads(html_info)
    totalPage = json_info['totalPage']
    project_list = []
    if totalPage:
        for i in range(int(totalPage)):
            params['queryPage'] = i + 1
            html_info = requests.get(url, params=params).content
            json_info = json.loads(html_info)
            project_list.extend(json_info['list'])
            print('爬取' + str(distStr) + '')
    for k, v in enumerate(project_list):
        Project_demonstration_level = get_project_details(v['PROJ_RID'])
        project_list[k]['Project_demonstration_level'] = Project_demonstration_level['Project_demonstration_level']
        project_list[k]['ESTIMATE_COPER'] = Project_demonstration_level['ESTIMATE_COPER']
        project_list[k]['Ways_of_purchasing_social_capital'] = Project_demonstration_level[
            'Ways_of_purchasing_social_capital']
        print('爬取' + str(distStr) + project_list[k]['PROJ_NAME'] + '详细信息')
    return project_list


def write_excel(title, distStr):
    if not os.path.exists('region.xlsx'):
        wb = openpyxl.Workbook()
        wb.save("region.xlsx")
    info = get_project_list_by_region(distStr)
    wb = openpyxl.load_workbook('region.xlsx')
    ws = wb.create_sheet(title=title)
    title = ['PROJ_NAME', 'PRV', 'IVALUE', 'INVESTCOUNT', 'PROJ_STATE_NAME', 'START_TIME',
             'Project_demonstration_level', 'RETURN_MODE_NAME', 'ESTIMATE_COPER', 'OPERATE_MODE_NAME',
             'Ways_of_purchasing_social_capital']
    ws.append(title)
    for i in info:
        cell_info = []
        for t in title:
            text = ILLEGAL_CHARACTERS_RE.sub(r'', str(i[t]))
            cell_info.append(text)
        print(cell_info)
        ws.append(cell_info)
    wb.save("region.xlsx")


if __name__ == '__main__':
    prv = get_region_and_filtvalue()
    print(prv)
    # prv_queue = Queue()
    #
    # for i in prv:
    #     prv_queue.put(i)
    # treading_list=[]
    # for i in range(8):
    #     treading_list.append(treading.Thread(target=write_excel(),args = ()))

    for p in prv:
        write_excel(p, prv[p])
